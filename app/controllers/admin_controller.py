import os
import pdfkit
import pandas as pd
from flask import render_template, request, redirect, url_for, send_file
from openpyxl import load_workbook
from config import Config

config = pdfkit.configuration(wkhtmltopdf=Config.WKHTMLTOPDF_PATH)
file_path = Config.EXCEL_PATH

# Ensure output directory
output_dir = os.path.join(os.path.dirname(__file__), "..", "static", "files")
os.makedirs(output_dir, exist_ok=True)
pdf_output_path = os.path.join(output_dir, "view.pdf")

def read_excel():
    """Read and return Excel data as a DataFrame."""
    df = pd.read_excel(file_path)
    return df

def update_excel(row, column, new_value):
    """Update a specific cell in the Excel file."""
    wb = load_workbook(file_path)
    sheet = wb.active
    sheet.cell(row=row, column=column, value=new_value)
    wb.save(file_path)

def generate_pdf():
    """Convert Excel data to HTML and then to PDF."""
    df = read_excel()
    html_content = render_template("pdf_template.html", table=df.to_html(classes="table table-bordered"))
    
    pdfkit.from_string(html_content, pdf_output_path, configuration=config)
    return pdf_output_path

def dashboard_controller():
   return render_template("admin/dashboard.html")


def update_controller():
    """Handle CRUD operations (update)."""
    row = int(request.form['row'])
    column = int(request.form['column'])
    new_value = request.form['new_value']

    update_excel(row, column, new_value)
    generate_pdf()  # Regenerate PDF after update

    return redirect(url_for('dashboard_controller'))


def pdf_templates():
    return render_template("file_layout/main_layout.html")